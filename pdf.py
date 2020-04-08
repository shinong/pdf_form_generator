from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import sqlite3
import csv
from datetime import (date,timedelta)
from math import exp,log
from openpyxl import load_workbook
import os

def pdfGen(run_num,input_list,dates,dpm):
    CellName_X = 95
    CellStep_Y = 20
    CellSize_Y = 25
    sample_list = input_list 

    packet = io.BytesIO()
    # create a new PDF with Reportlab
    can = canvas.Canvas(packet, pagesize=letter)
    for i in range(CellSize_Y):
        can.drawString(CellName_X, 80+i*20, str(sample_list[24-i]))
    can.drawString(180,615,'{}'.format(dates[1]))                                          #put 'end_date' here later 
    can.drawString(180,630,'{}'.format(dates[0]))                                          #put 'start_date' here later
    can.drawString(235,650,'118')
    can.drawString(245,675,'Shinong Mao')
    can.drawString(360,615,'1.0')
    can.drawString(360,630,'250')
    can.drawString(360,645,'5.8')
    can.drawString(480,615,'678.6')
    can.drawString(480,630,'{}'.format(round(dpm,3)))
    can.drawString(485,675,run_num)
    can.save()

    #move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader(open("template.pdf", "rb"))
    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("destination.pdf", "wb")
    output.write(outputStream)
    outputStream.close()

def pdfWeightGen(weightList):
    cellSize_Y = 25
    cellSize_X = 3
    cellStep_Y = 19.9
    cellStep_X = 65
    cellStart_X = 225
    cellStart_Y = 81

    packet = io.BytesIO()
    # create a new PDF with Reportlab
    can = canvas.Canvas(packet, pagesize=letter)
    
    for i in range(cellSize_X):
        for j in range(cellSize_Y):
            can.drawString(cellStart_X+i*cellStep_X,cellStart_Y+j*cellStep_Y,str(i)+str(j))
    can.save()
    #move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader(open("template.pdf", "rb"))
    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("destination.pdf", "wb")
    output.write(outputStream)
    outputStream.close()
    

def listPGen(E1,Spk,S_list):
    if len(S_list) != 20:
        print('error, number does not match!')
        return 0
    new_E1 = (E1 + 3)%25
    new_Spk = (Spk + 1)%25
    if new_Spk == new_E1:               #overlap with E1
        new_E1 = (new_E1 + 1)%25
        new_E2 = (new_E1 + 1)%25
        new_E3 = (new_E1 + 2)%25
        new_DI = (new_E1 + 3)%25
    elif new_Spk == new_E1 +1:          #overlap with E2
        new_E2 = (new_E1 + 2)%25
        new_E3 = (new_E1 + 3)%25
        new_DI = (new_E1 + 4)%25
    elif new_Spk == new_E1 +2:          #overlap with E3
        new_E2 = (new_E1 + 1)%25
        new_E3 = (new_E1 + 3)%25
        new_DI = (new_E1 + 4)%25
    elif new_Spk == new_E1 +3:          #overlap with DI
        new_E2 = (new_E1 + 1)%25
        new_E3 = (new_E1 + 2)%25
        new_DI = (new_E1 + 4)%25
    else:                               #no overlap
        new_E2 = (new_E1 + 1)%25
        new_E3 = (new_E1 + 2)%25
        new_DI = (new_E1 + 3)%25

    if new_E1 == 0:
        new_E1 = 25
    elif new_E2 == 0:
        new_E2 = 25
    elif new_E3 == 0:
        new_E3 = 25
    elif new_Spk == 0:
        new_Spk = 25
    elif new_DI == 0:
        new_DI = 25
    else:
        pass
    
    new_list = []
    counter = 0
    for i in range(25):
        if i+1 == new_E1 or i+1 == new_E2 or i+1 == new_E3:
            new_list.append('EBKG')
            counter = counter+1
        elif i+1 == new_Spk:
            new_list.append('SPIKE')
            counter = counter+1
        elif i+1 == new_DI:
            new_list.append('DI')
            counter = counter+1
        else:
            new_list.append(S_list[i-counter])
            pass
        
    return new_list

def listJGen(E1,Spk,S_list):      #number 6 and 13 are fixed 
    if len(S_list) != 19:
        print('error, number does not match!')
        return 0
    
    if E1 >6 and E1 < 13:               #the input number is the posion in 25 samples, need to be renormalized
        E1 = E1 -1 
    elif E1 > 13:
        E1 = E1 -2
    if Spk >6 and Spk < 13:              
        Spk = Spk -1 
    elif Spk > 13:
        Spk = Spk -2
    
    new_E1 = (E1 + 3)%23
    new_Spk = (Spk + 1)%23
    print(new_E1,new_Spk)

    if new_Spk == new_E1:               #overlap with E1
        new_E1 = (new_E1 + 1)%23
        new_E2 = (new_E1 + 1)%23
        new_E3 = (new_E1 + 2)%23
    elif new_Spk == new_E1 +1:          #overlap with E2
        new_E2 = (new_E1 + 2)%23
        new_E3 = (new_E1 + 3)%23
    elif new_Spk == new_E1 +2:          #overlap with E3
        new_E2 = (new_E1 + 1)%23
        new_E3 = (new_E1 + 3)%23
    elif new_Spk == new_E1 +3:          #overlap with DI
        new_E2 = (new_E1 + 1)%23
        new_E3 = (new_E1 + 2)%23
    else:                               #no overlap
        new_E2 = (new_E1 + 1)%23
        new_E3 = (new_E1 + 2)%23
    

    if new_E1 == 0:
        new_E1 = 23
    elif new_E2 == 0:
        new_E2 = 23
    elif new_E3 == 0:
        new_E3 = 23
    elif new_Spk == 0:
        new_Spk = 23
    else:
        pass

    new_list = []
    counter = 0
    for i in range(23):
        if i+1 == new_E1 or i+1 == new_E2 or i+1 == new_E3:       
            new_list.append('EBKG')
            counter = counter+1
        elif i+1 == new_Spk:
            new_list.append('SPIKE')
            counter = counter+1
        else:
            new_list.append(S_list[i-counter])
            pass
    new_list.insert(5,'EDI-Prime')              #insert the 2 cell in the list
    new_list.insert(12,'EDI-Prime')

    return new_list



def readDb(table):
    connection = sqlite3.connect('sampleform.db')
    cursor = connection.cursor()
    

def testDbGen():
    pass

def csvReader(path):
    new_list = []
    f = open(path,'r', newline = '')
    reader = csv.reader(f)
    for i in reader:
        new_list.append(i[0])
    f.close()
    return new_list

def csvWriter(path,sample_list,runNum,dpm):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    tempfilename = os.path.join(fileDir, '../E3H/template.xlsx')
    tempfilename = os.path.abspath(os.path.realpath(tempfilename))
    savefilename = os.path.join(fileDir, path)
    savefilename = os.path.abspath(os.path.realpath(savefilename))
    wb = load_workbook(tempfilename)
    sheet = wb['Input']
    for i in range(len(sample_list)):
        sheet.cell(row=3+i,column = 5).value = sample_list[i]
    sheet.cell(row = 2, column = 3).value = runNum
    sheet.cell(row = 30, column = 3).value = dpm
    wb.save(savefilename)
    
'''
def csvWriter(path,sample_list):
    f = open(path,'w',newline = '')
    writer = csv.writer(f)
    print(len(sample_list))
    for i in range(len(sample_list)):
        writer.writerow([sample_list[i]])
    f.close()
'''
def timeCal(QtDate):
    tempList = QtDate.split('/')
    y,m,d = int(tempList[2]),int(tempList[0]),int(tempList[1])
    start_date = date(y,m,d)
    interval_days = 5
    end_date = start_date + timedelta(days = interval_days)
    return [start_date,end_date]

def dpmCal(y,m):
    T10 = 47.919
    startDate = date(2019,5,1)
    currentDate = date(y,m,1)
    days = currentDate - startDate
    return T10*exp(-1*log(2)/4500*days.days)*8


if __name__ == "__main__":
    #sample_list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25]
    #pdfGen(sample_list)
    sampleP_list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20]
    E= 16
    SPK = 1
    #sampleP_list = csvReader('list_input.csv')
    #print(listPGen(E,SPK,sampleP_list))
    #print(listJGen(E,SPK,sampleP_list))
    csvWriter('../E3H/PRun/Run {}.xlsx',listPGen(E,SPK,sampleP_list),'p123',123.45)
    #csvWriter('list_output.csv',listJGen(E,SPK,sampleP_list))
    #pdfGen('P660',listPGen(E,SPK,sampleP_list))
    #pdfGen('1428',listJGen(E,SPK,sampleP_list),timeCal(2019,8,22))
    #print(timeCal('9/10/2019'))
    print(dpmCal(2019,8))
    print(dpmCal(2019,9))
    print(dpmCal(2019,10))
    print(dpmCal(2019,11))
    print(dpmCal(2019,12))
    print(dpmCal(2020,8))
    pdfWeightGen(8)
